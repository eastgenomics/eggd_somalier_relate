import pandas as pd
import argparse
from openpyxl import load_workbook

## To do: Filenaming system !! How to get project filename? dx pwd??

def parse_args():
    """
    Allow arguments from the command line to be given. Array of somalier files is needed but 
    Reported sex file is not always required
    """
    ## Read in arguments
    # The ArgumentParser object will hold all the information necessary to parse the command line into Python data types.
    parser = argparse.ArgumentParser(description='Process an array of somalier files.')

    # Filling an ArgumentParser with information about program arguments is done by making calls to the add_argument() method. 
    # Generally, these calls tell the ArgumentParser how to take the strings on the command line and turn them into objects. 
    parser.add_argument('-a', '--array', nargs='+', required=True, help='an array of somalier files')

    parser.add_argument('-s', '--reported_sex', nargs='+', required=False, help='csv/tsv/xlsx of known sample sex')

    args = parser.parse_args()

    print("Inputs are: " , args.array)

    return args

def get_sampleID(args):
    """
    Gets the sampleID from the {sample}.somalier files inputted into app.
    """
    samplesID = []
   
    for a in args.array:
        samplesID.append(a.split(".")[0])
    
    return samplesID

def read_col_LABNO(sheet, max_row):
    """
    Reads in LABNO column of the xlsx which is in the first column
    """
    LABNO = []
    for i in range(1, max_row+1):
        # get particular cell value    
        cell_obj=sheet.cell(row=i,column=1)
        # print cell value     
        LABNO.append(cell_obj.value)
        
    return LABNO

def read_col_ExomeNumber(sheet, max_row):
    """
    Reads in ExomeNumber column of the xlsx which is in the second column
    """
    ExomeNumber = []
    for i in range(1, max_row+1):
        # get particular cell value    
        cell_obj=sheet.cell(row=i,column=2)
        # print cell value     
        ExomeNumber.append(cell_obj.value)
        
    return ExomeNumber

def read_col_SEX(sheet, max_row):
    """
    Reads in sex column of the xlsx which is in the third column
    """
    SEX = []
    for i in range(1, max_row+1):
        # get particular cell value    
        cell_obj=sheet.cell(row=i,column=3)
        # print cell value     
        SEX.append(cell_obj.value)
        
    return SEX

def known_sex(samplesID, args):
    """
    Produce a ped file where all sex is known from xlsx file where sample's sex is listed. 
    Assumes input is xlsx and list of sample names.
    """
    file = args.reported_sex[0]
    dfs = load_workbook(filename = file) #load xlsx #https://medium.com/aubergine-solutions/working-with-excel-sheets-in-python-using-openpyxl-4f9fd32de87f
    sheet=dfs.active #select the xlsx
    max_row=sheet.max_row # get max row count

    labno = read_col_LABNO(sheet,max_row)
    exomenumber = read_col_ExomeNumber(sheet,max_row)
    sex = read_col_SEX(sheet, max_row)

    ReportedSex = []
    # For every sample is list, if sample exists in exomenumber, pull out the index and select their sex at that index. 
    # If sample not in list, then provide unknown
    for i in range(0,len(samplesID)):
        sample = samplesID[i]
        if sample in exomenumber:
            index = exomenumber.index(sample)
            ReportedSex.append(sex[index])

        else:
            print(sample, " does not exist in Gemini db")
            #Append unknown to the reportedSex
            ReportedSex.append("U")

    print("---- Reported Sex -----")
    print(samplesID)
    print(ReportedSex)

    ## Ped uses number instead of F/M So replace letter with number. Sex code ('1' = male, '2' = female, '0' = unknown)
    ReportedSex = [s.replace('M','1') for s in ReportedSex]
    ReportedSex = [s.replace('F','2') for s in ReportedSex]
    ReportedSex = [s.replace('U','0') for s in ReportedSex]
    print(ReportedSex)

    FamilyID = samplesID
    PaternalID = [0] * len(samplesID)
    MaternalID = [0] * len(samplesID)
    Sex = ReportedSex
    Phenotype = [-9] * len(samplesID)

    print("--------------Making PED FILE-----------")
    df = pd.DataFrame(list(zip(FamilyID, samplesID,  PaternalID, MaternalID, Sex, Phenotype)), columns =['FID', 'IID', 'PaternalID', 'MaternalID','Sex', 'Phenotype'])
    print(df)

    df.to_csv('Samples.ped', sep='\t',index=False, header =False)

def unknown_sex(samplesID):
    """
    Produce a ped file where all sex is unknown
    """
    FamilyID = samplesID
    PaternalID = [0] * len(samplesID)
    MaternalID = [0] * len(samplesID)
    Sex = [0] * len(samplesID) #mark all samples sex as 0 sinc they're unknown 
    Phenotype = [-9] * len(samplesID)

    print("--------------Making PED FILE-----------")
    df = pd.DataFrame(list(zip(FamilyID, samplesID,  PaternalID, MaternalID, Sex, Phenotype)), columns =['FID', 'IID', 'PaternalID', 'MaternalID','Sex', 'Phenotype'])
    print(df)

    df.to_csv('Samples.ped', sep='\t',index=False, header =False)

def main():
    """
    Main function to produce input sample's ped file.
    """
    args = parse_args()
    samplesID = get_sampleID(args)

    if args.reported_sex == None:
        print("Reported sex file is not provided - all sample's sex are unknown.")
        unknown_sex(samplesID)
    else: 
        print("Reported sex file is provided.")
        known_sex(samplesID,args)

if __name__ == "__main__":

    main()
